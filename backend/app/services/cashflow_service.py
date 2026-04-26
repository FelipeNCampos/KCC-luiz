from io import BytesIO
from datetime import date
from decimal import Decimal

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.cashflow import CashFlowRecord
from app.repositories.cashflow_repository import CashFlowRepository
from app.schemas.cashflow import CashFlowCreate, CashFlowListResponse, CashFlowRow
from app.services.email_service import EmailService


class CashFlowService:
    def __init__(self, repository: CashFlowRepository, email_service: EmailService | None = None) -> None:
        self.repository = repository
        self.email_service = email_service or EmailService()

    def create_record(
        self,
        user_id: int,
        payload: CashFlowCreate,
        invoice_media_name: str | None = None,
        invoice_media_mime: str | None = None,
        invoice_media_data: bytes | None = None,
    ) -> CashFlowRecord:
        if payload.has_invoice and not invoice_media_data:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Invoice media is required when invoice is Yes",
            )

        amount = abs(payload.value)
        if payload.entry_type == "outcome":
            amount = -amount

        record = CashFlowRecord(
            payment_number=self.repository.get_next_payment_number(),
            has_invoice=payload.has_invoice,
            invoice_media_name=invoice_media_name if payload.has_invoice else None,
            invoice_media_mime=invoice_media_mime if payload.has_invoice else None,
            invoice_media_data=invoice_media_data if payload.has_invoice else None,
            record_date=payload.record_date,
            amount=amount,
            description=payload.description.strip(),
            flat=payload.flat.strip(),
            created_by_user_id=user_id,
        )
        return self.repository.create(record)

    def list_month(self, month: str | None, search: str | None = None) -> CashFlowListResponse:
        month_label, month_start, month_end = self._parse_month(month)
        records = self.repository.list_month_records(month_start, month_end)

        query = (search or "").strip().lower()
        running_balance = Decimal("0")
        monthly_total = Decimal("0")
        items: list[CashFlowRow] = []

        for record in records:
            monthly_total += record.amount
            running_balance += record.amount

            if query and query not in record.description.lower() and query not in record.flat.lower():
                continue

            items.append(
                CashFlowRow(
                    id=record.id,
                    payment_number=record.payment_number,
                    has_invoice=record.has_invoice,
                    invoice_media_name=record.invoice_media_name,
                    record_date=record.record_date,
                    amount=record.amount,
                    description=record.description,
                    flat=record.flat,
                    balance=running_balance,
                    created_by_user_id=record.created_by_user_id,
                    created_at=record.created_at,
                )
            )

        return CashFlowListResponse(month=month_label, monthly_total=monthly_total, items=items)

    def delete_record(self, record_id: int) -> None:
        record = self.repository.get_by_id(record_id)
        if record is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cash flow record not found")
        self.repository.delete(record)

    def get_invoice_media(self, record_id: int) -> tuple[str, str, bytes]:
        record = self.repository.get_by_id(record_id)
        if record is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cash flow record not found")
        if not record.has_invoice or not record.invoice_media_data or not record.invoice_media_mime:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invoice media not found")

        return (
            record.invoice_media_name or "invoice",
            record.invoice_media_mime,
            record.invoice_media_data,
        )

    def send_month_report(self, recipient: str, month: str, search: str | None = None) -> None:
        listing = self.list_month(month=month, search=search)
        _, month_start, _ = self._parse_month(month)
        opening_balance = self.repository.get_balance_before(month_start)
        closing_balance = opening_balance + listing.monthly_total
        spreadsheet_data = self._build_report_workbook(listing, opening_balance, closing_balance, search)
        file_name = f"cashflow-report-{listing.month}.xlsx"
        subject = f"Cashflow report {listing.month}"
        body = f"Attached is the cashflow report for {listing.month}."

        self.email_service.send_report(
            recipient=recipient,
            subject=subject,
            body=body,
            attachment_name=file_name,
            attachment_data=spreadsheet_data,
        )

    @staticmethod
    def _build_report_workbook(
        listing: CashFlowListResponse,
        opening_balance: Decimal,
        closing_balance: Decimal,
        search: str | None,
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Cashflow Report"

        header_fill = PatternFill("solid", fgColor="8C7569")
        header_font = Font(color="FFFFFF", bold=True)
        section_fill = PatternFill("solid", fgColor="F5F1EE")
        title_font = Font(size=14, bold=True, color="55311C")
        bold_font = Font(bold=True, color="55311C")
        border = Border(
            left=Side(style="thin", color="E5E0DC"),
            right=Side(style="thin", color="E5E0DC"),
            top=Side(style="thin", color="E5E0DC"),
            bottom=Side(style="thin", color="E5E0DC"),
        )
        currency_format = '$#,##0.00;[Red]-$#,##0.00'

        sheet["A1"] = "Cashflow Report"
        sheet["A1"].font = title_font
        sheet["A2"] = f"Month: {listing.month}"
        sheet["A3"] = f"Filter: {search.strip()}" if search and search.strip() else "Filter: All records"

        summary_rows = [
            ("Opening Balance", opening_balance),
            ("Monthly Balance", listing.monthly_total),
            ("Closing Balance", closing_balance),
        ]

        for index, (label, value) in enumerate(summary_rows, start=5):
            sheet[f"A{index}"] = label
            sheet[f"A{index}"].font = bold_font
            sheet[f"A{index}"].fill = section_fill
            sheet[f"A{index}"].border = border
            sheet[f"B{index}"] = float(value)
            sheet[f"B{index}"].number_format = currency_format
            sheet[f"B{index}"].font = bold_font
            sheet[f"B{index}"].fill = section_fill
            sheet[f"B{index}"].border = border

        start_row = 10
        headers = ["Payment Number", "Invoice", "Date", "Amount", "Description", "Flat", "Balance"]
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_index, item in enumerate(listing.items, start=start_row + 1):
            values = [
                item.payment_number,
                "Yes" if item.has_invoice else "No",
                item.record_date.isoformat(),
                float(item.amount),
                item.description,
                item.flat,
                float(item.balance),
            ]
            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if row_index % 2 == 0:
                    cell.fill = section_fill
                if column_index in {4, 7}:
                    cell.number_format = currency_format

        widths = {
            1: 18,
            2: 14,
            3: 14,
            4: 14,
            5: 36,
            6: 14,
            7: 14,
        }
        for column_index, width in widths.items():
            sheet.column_dimensions[get_column_letter(column_index)].width = width

        sheet.freeze_panes = "A11"

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _parse_month(month: str | None) -> tuple[str, date, date]:
        if month is None or not month.strip():
            today = date.today()
            month_start = date(today.year, today.month, 1)
            month_label = f"{today.year:04d}-{today.month:02d}"
        else:
            try:
                year_str, month_str = month.split("-", maxsplit=1)
                year = int(year_str)
                month_value = int(month_str)
                month_start = date(year, month_value, 1)
                month_label = f"{year:04d}-{month_value:02d}"
            except (TypeError, ValueError) as exc:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Invalid month format. Use YYYY-MM",
                ) from exc

        if month_start.month == 12:
            month_end = date(month_start.year + 1, 1, 1)
        else:
            month_end = date(month_start.year, month_start.month + 1, 1)

        return month_label, month_start, month_end
